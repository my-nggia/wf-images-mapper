import os
import io
import zipfile
import os
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import json

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# Sheets that are NOT product sheets
NON_PRODUCT_SHEETS = {
    'Instructions', 'WAYFAIR_USE_ONLY', 'Valid Values',
    'Additional Cartons', 'Additional Images', 'Additional Videos',
    'Additional Documents', 'Additional Chemicals', 'Failed Products'
}

# How many header rows before actual data rows
HEADER_ROW_COUNT = 8  # rows 0-7 are headers/meta, data starts at row index 8


def get_product_sheets(wb):
    return [s for s in wb.sheetnames if s not in NON_PRODUCT_SHEETS]


def extract_skus_from_template(xlsx_bytes):
    """Extract all SKUs from all product sheets in the template."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    product_sheets = get_product_sheets(wb)
    
    skus = {}  # sku -> sheet_name
    for sheet_name in product_sheets:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < HEADER_ROW_COUNT:
                continue
            sku = row[0] if row[0] else None
            if sku and str(sku).strip():
                skus[str(sku).strip()] = sheet_name
    return skus, product_sheets


def build_excel_output(xlsx_bytes, image_map):
    """
    Create one XLSX file with 2 sheets:
    - Main Images
    - Additional Images
    """

    wb = Workbook()

    # Sheet 1
    ws_main = wb.active
    ws_main.title = "Main Images"

    main_headers = [
        'Supplier Part Number',
        'Image File Name or URL 1',
        'Image File Name or URL 2',
        'Image File Name or URL 3',
        'Image File Name or URL 4',
        'Image File Name or URL 5'
    ]

    ws_main.append(main_headers)

    # Sheet 2
    ws_add = wb.create_sheet("Additional Images")

    add_headers = [
        'Supplier Part Number',
        'Image File Name or URL'
    ]

    ws_add.append(add_headers)

    for sku, images in sorted(image_map.items()):

        # Main sheet
        main_row = [sku]

        for i in range(5):
            if i < len(images):
                main_row.append(images[i])
            else:
                main_row.append('')

        ws_main.append(main_row)

        # Additional Images sheet
        for img in images[5:]:
            ws_add.append([sku, img])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/process', methods=['POST'])
def process():
    print("FILES:", list(request.files.keys()))
    print("FORM:", list(request.form.keys()))
    try:
        # Get template file
        if 'template' not in request.files:
            return jsonify({'error': 'Thiếu file Template (.xlsx)'}), 400
        
        template_file = request.files['template']
        if not template_file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Template phải là file .xlsx'}), 400

        xlsx_bytes = template_file.read()

        # Get image files
        image_names_json = request.form.get('image_names')

        if not image_names_json:
            return jsonify({
                'error': 'Không nhận được danh sách ảnh'
            }), 400

        image_names = json.loads(image_names_json)

        # Extract valid image extensions
        IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tiff', '.tif'}

        # Extract SKUs from template
        skus_in_template, product_sheets = extract_skus_from_template(xlsx_bytes)

        # Build image map: SKU -> [filenames]
        # Image filename format: SKU_something.ext or SKU.ext or SKUxxx.ext
        # Strategy: filename stem starts with SKU (case-insensitive)
        image_map = defaultdict(list)
        unmatched = []
        matched_count = 0

        # Sort images for consistent ordering
        image_map = defaultdict(list)
        unmatched = []
        matched_count = 0

        sorted_images = sorted(image_names, key=str.lower)

        for original_name in sorted_images:

            name_stem, ext = os.path.splitext(original_name)

            if ext.lower() not in IMAGE_EXTS:
                continue

            matched_sku = None
            matched_len = 0

            for sku in skus_in_template:

                if name_stem.upper().startswith(sku.upper()):

                    if len(sku) > matched_len:
                        matched_sku = sku
                        matched_len = len(sku)

            if matched_sku:
                image_map[matched_sku].append(original_name)
                matched_count += 1
            else:
                unmatched.append(original_name)

        # Build CSVs
        # Build Excel file
        excel_buffer = build_excel_output(xlsx_bytes, image_map)
        app.config['_last_excel'] = excel_buffer.getvalue()        

        # Summary stats
        skus_matched = len(image_map)
        skus_with_extra = sum(1 for imgs in image_map.values() if len(imgs) > 5)
        total_images = sum(len(imgs) for imgs in image_map.values())

        app.config['_last_excel'] = excel_buffer.getvalue()

        app.config['_last_has_additional'] = any(
            len(images) > 5
            for images in image_map.values()
        )

        return jsonify({
            'success': True,
            'stats': {
                'skus_in_template': len(skus_in_template),
                'skus_matched': skus_matched,
                'total_images': total_images,
                'skus_with_extra': skus_with_extra,
                'unmatched_count': len(unmatched),
                'unmatched_sample': unmatched[:10],
                'product_sheets': product_sheets,
            },
            'has_additional': app.config['_last_has_additional']
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/download')
def download():
    excel_data = app.config.get('_last_excel')

    if not excel_data:
        return 'No file ready', 404

    return send_file(
        io.BytesIO(excel_data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='wayfair_image_mapping.xlsx'
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
